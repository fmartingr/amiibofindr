# coding: utf-8

# python3
from __future__ import unicode_literals

# python
from pprint import pprint
import re
import uuid

# third party
from openpyxl import load_workbook
import requests

# django
from django.core.management.base import BaseCommand, CommandError
from django.utils import translation
from django.utils.text import slugify

# amiibofindr
from amiibofindr.apps.amiibo.models import Amiibo, Collection, AmiiboShop
from amiibofindr.apps.shop.models import Shop


class AmiiboSheet(object):
    DOC_ID = '1uOZKhVW1S25GDETsqinf0Qzl1trbKcw2wT2swfcTAfE'

    SHEET_NA_US = 'NA (US)'
    SHEET_NA_CA = 'NA (CA)'
    SHEET_EU = 'EU'
    SHEET_JP = 'JP'

    SERIES_STRING_US = 'series'
    SERIES_STRING_EU = 'Collection'
    SERIES_STRING_JP = 'シリーズ'

    LINK_REGEX = r'\=HYPERLINK\(\"(.*)\",\"(.*)\"\)'

    SHEETS = {
        SHEET_NA_US: {
            'SERIES_STRING': SERIES_STRING_US,
            'SHOP_HEADS': {
                u'Amazon ASIN': u'amazon-us',
            },
        },
        # SHEET_NA_CA: {
        #     'SERIES_STRING': SERIES_STRING,
        # },
        SHEET_EU: {
            'SERIES_STRING': SERIES_STRING_EU,
            'SHOP_HEADS': {
                u'UK ASIN': u'amazon-uk',
                u'FR ASIN': u'amazon-fr',
                u'ES ASIN': u'amazon-es',
                u'DE ASIN': u'amazon-de',
                u'IT ASIN': u'amazon-it',
            },
        },
        SHEET_JP: {
            'SERIES_STRING': SERIES_STRING_JP,
            'SHOP_HEADS': {
                u'Amazon ASIN': u'amazon-jp',
            },
        },
    }

    DATA = {}

    def __init__(self):
        self.file = '/tmp/{}.xlsx'.format(uuid.uuid4())
        self.download_xlsx()
        self.workbook = load_workbook(self.file)

    def download_xlsx(self):
        r = requests.get(
            'https://docs.google.com/spreadsheets/d/{}/export?format=xlsx'.format(
                self.DOC_ID))

        with open(self.file, 'wb') as fd:
            for chunk in r.iter_content(255):
                fd.write(chunk)

    def parse(self):
        for sheet in self.SHEETS:
            sheet_conf = self.SHEETS[sheet]
            ws = self.workbook[sheet]
            current_series = ''
            self.DATA[ws.title] = []
            first = True

            for row in ws.iter_rows():
                if first:
                    first = False
                    continue

                amiibo = {}
                for cell in row:
                    if cell.internal_value:
                        # Discard bundles
                        #if cell.column == 'A' and cell.internal_value == 'BUNDLE':
                        #    break;

                        if (
                            sheet_conf['SERIES_STRING'] in unicode(cell.value)
                            and cell.column == 'A'
                        ):
                            current_series = cell.value.replace(
                                sheet_conf['SERIES_STRING'], ''
                            ).strip()
                            continue
                        else:
                            amiibo['collection'] = current_series

                            cell_head = ws['{}1'.format(cell.column)].value

                            if getattr(cell, 'data_type', None) == 'f':
                                matches = re.match(
                                    self.LINK_REGEX, cell.internal_value)
                                if matches:
                                    cell_name = 'id'
                                    if cell_head == 'amiibo':
                                        cell_name = 'name'

                                    amiibo[cell_head] = {
                                        'link': matches.group(1),
                                        cell_name: matches.group(2),
                                    }
                            else:
                                amiibo[cell_head] = unicode(
                                    cell.internal_value).replace('.0', '')

                if amiibo.keys():
                    self.DATA[ws.title].append(amiibo)


class Command(BaseCommand):
    def handle(self, *args, **options):
        self.xls_parser = AmiiboSheet()
        self.xls_parser.parse()

        self.parse_data(self.xls_parser.DATA)

    def parse_data(self, data):
        for sheet in self.xls_parser.SHEETS:
            sheet_obj = self.xls_parser.SHEETS[sheet]
            print('==> SHEET: {}'.format(sheet))
            for amiibo in data[sheet]:
                pprint(amiibo)

                if 'Model No.' in amiibo:
                    try:
                        amiibo_obj = Amiibo.objects.get(
                            model_number=amiibo['Model No.'])
                        created = False
                    except Amiibo.DoesNotExist:
                        created = True
                        amiibo_obj = Amiibo(model_number=amiibo['Model No.'])

                    if created:
                        collection_obj, new = Collection.objects.get_or_create(
                            name_eu=amiibo['collection']
                        )

                        amiibo_obj.collection = collection_obj

                    if 'amiibo' in amiibo:
                        if isinstance(amiibo['amiibo'], unicode):
                            amiibo_name = amiibo['amiibo']
                            amiibo_link = ''
                        else:
                            amiibo_name = amiibo['amiibo']['name']
                            amiibo_link = amiibo['amiibo']['link']

                        if sheet == self.xls_parser.SHEET_NA_US:
                            amiibo_obj.name_us = amiibo_name
                            amiibo_obj.link_us = amiibo_link
                            amiibo_obj.collection.name_us = amiibo['collection']
                            amiibo_obj.collection.save()

                        if sheet == self.xls_parser.SHEET_EU:
                            amiibo_obj.name_eu = amiibo_name
                            amiibo_obj.link_eu = amiibo_link

                        if sheet == self.xls_parser.SHEET_JP:
                            amiibo_obj.name_jp = amiibo_name
                            amiibo_obj.link_jp = amiibo_link
                            amiibo_obj.collection.name_jp = amiibo['collection']
                            amiibo_obj.collection.save()

                        if amiibo_obj.name_eu and not amiibo_obj.slug:
                            amiibo_obj.slug = slugify(amiibo_obj.name_eu)

                    try:
                        amiibo_obj.collection_number = int(amiibo['#'])
                    except:
                        amiibo_obj.collection_number = None

                    amiibo_obj.save()

                    for shop_head, shop_slug in sheet_obj['SHOP_HEADS'].iteritems():
                        if shop_head in amiibo:
                            try:
                                shop = Shop.objects.get(slug=shop_slug)
                                amiibo_shop, created = AmiiboShop.objects.get_or_create(
                                    amiibo_id=amiibo_obj.pk,
                                    shop_id=shop.pk
                                )
                                amiibo_shop.url = amiibo[shop_head]['link']
                                amiibo_shop.save()
                            except Shop.DoesNotExist:
                                pass
